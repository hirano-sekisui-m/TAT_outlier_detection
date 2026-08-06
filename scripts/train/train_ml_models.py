#!/usr/bin/env python3
"""
Flexible Machine Learning Model Training & Evaluation Script.

Supports both Classification and Regression tasks.
Evaluates multiple algorithms via Cross-Validation and generates comprehensive summary reports.
"""

import argparse
import json
import platform
import warnings
from datetime import datetime, timezone
from pathlib import Path

import joblib
import numpy as np
import pandas as pd

# Scikit-learn imports
import sklearn
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from sklearn.ensemble import (
    ExtraTreesClassifier,
    ExtraTreesRegressor,
    GradientBoostingClassifier,
    GradientBoostingRegressor,
    RandomForestClassifier,
    RandomForestRegressor,
)
from sklearn.linear_model import Lasso, LogisticRegression, Ridge
from sklearn.metrics import (
    accuracy_score,
    average_precision_score,
    confusion_matrix,
    f1_score,
    mean_absolute_error,
    mean_squared_error,
    precision_score,
    r2_score,
    recall_score,
    roc_auc_score,
)
from sklearn.model_selection import KFold, StratifiedKFold, cross_val_predict
from sklearn.pipeline import make_pipeline
from sklearn.preprocessing import StandardScaler
from sklearn.svm import SVC, SVR

warnings.filterwarnings("ignore")

# プロジェクトルートディレクトリの設定
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent


# ==========================================================
# 設定エリア (パスや学習条件を変更したい場合はここを修正してください)
# ==========================================================

# 1. 入力・出力パスの設定
# 入力データセット: build_ml_dataset.py で出力されたExcelファイル
DEFAULT_INPUT_DATASET = PROJECT_ROOT / "data" / "260806_TAT1_dataset.xlsx"

# サマリーレポートの出力先フォルダ
DEFAULT_REPORT_DIR = PROJECT_ROOT / "reports" / "train" / "260806_TAT1_ml_training"

# 学習済みモデルの保存先フォルダ
DEFAULT_MODEL_DIR = PROJECT_ROOT / "models" / "trained" /"260806_TAT1_ml_models"


# 2. タスク種別の選択: "classification" (分類) または "regression" (回帰)
TASK_TYPE = "classification"


# 3. 正解ラベル (Target) の動的定義関数
def define_target_label(df: pd.DataFrame, task_type: str = TASK_TYPE) -> tuple[pd.Series, str]:
    """
    Pythonスクリプト内で正解ラベル(目的変数)を柔軟に定義します。
    (ターゲット要素の Series, 正解ラベルの定義説明テキスト) のタプルを返します。
    
    【分類タスク (classification) の例】
      - 14-20/2-8 (終点/始点比率) が 1.15 を超える場合を 1 (高高値/高乖離陽性)、それ以外を 0
      
    【回帰タスク (regression) の例】
      - TAT1装置生データ、または特定の数値カラムを連続数値として予測対象にする
    """
    if task_type == "classification":
        if "205/206" in df.columns:
            val = pd.to_numeric(df["205/206"], errors="coerce")
            target = (val > 1.15).astype(int)
            target_def = "205/206 > 1.15"
        elif "終点/始点" in df.columns:
            val = pd.to_numeric(df["終点/始点"], errors="coerce")
            target = (val > 1.15).astype(int)
            target_def = "終点/始点 > 1.15"
        else:
            val = pd.to_numeric(df.get("TAT1装置生データ", 0), errors="coerce")
            target = (val > 5.0).astype(int)
            target_def = "TAT1装置生データ > 5.0"
        return target, target_def

    elif task_type == "regression":
        target = pd.to_numeric(df["205/206"], errors="coerce")
        target_def = "205/206"
        return target, target_def

    else:
        raise ValueError(f"未知のタスク種別です: {task_type}. 'classification' または 'regression' を指定してください。")


# 4. 学習に使用する特徴量 (説明変数) の定義
FEATURE_COLS = [
    "装置生データTAT1", "2-8", "4-10", "6-12", "8-14", "10-16", "12-18", "14-20",
    "平均", "最大", "最小", "レンジ", "最大上昇率", "最大上昇速度", "最大下落率", "最大下落速度",
    "4-10/2-8", "6-12/2-8", "8-14/2-8", "10-16/2-8", "12-18/2-8", "14-20/2-8"
]


# 5. その他のパラメータ
RANDOM_STATE = 42
N_SPLITS = 5

# ==========================================================


def get_classification_models(random_state: int = RANDOM_STATE) -> dict:
    """分類モデルの候補一覧を取得"""
    return {
        "RandomForest": RandomForestClassifier(n_estimators=500, min_samples_leaf=3, class_weight="balanced", random_state=random_state, n_jobs=-1),
        "GradientBoosting": GradientBoostingClassifier(n_estimators=200, random_state=random_state),
        "ExtraTrees": ExtraTreesClassifier(n_estimators=500, min_samples_leaf=3, class_weight="balanced", random_state=random_state, n_jobs=-1),
        "LogisticRegression": make_pipeline(StandardScaler(), LogisticRegression(class_weight="balanced", max_iter=1000, random_state=random_state)),
        "SVC": make_pipeline(StandardScaler(), SVC(probability=True, class_weight="balanced", random_state=random_state)),
    }


def get_regression_models(random_state: int = RANDOM_STATE) -> dict:
    """回帰モデルの候補一覧を取得"""
    return {
        "RandomForestRegressor": RandomForestRegressor(n_estimators=500, min_samples_leaf=3, random_state=random_state, n_jobs=-1),
        "GradientBoostingRegressor": GradientBoostingRegressor(n_estimators=200, random_state=random_state),
        "ExtraTreesRegressor": ExtraTreesRegressor(n_estimators=500, min_samples_leaf=3, random_state=random_state, n_jobs=-1),
        "Ridge": make_pipeline(StandardScaler(), Ridge(random_state=random_state)),
        "Lasso": make_pipeline(StandardScaler(), Lasso(alpha=0.1, random_state=random_state)),
        "SVR": make_pipeline(StandardScaler(), SVR()),
    }


def evaluate_classification_model(model_name: str, model, X: pd.DataFrame, y: pd.Series, cv):
    """分類モデルの交差検証評価"""
    try:
        y_prob = cross_val_predict(model, X, y, cv=cv, method="predict_proba")[:, 1]
    except (AttributeError, NotImplementedError):
        y_prob = cross_val_predict(model, X, y, cv=cv, method="decision_function")
        y_prob = (y_prob - y_prob.min()) / (y_prob.max() - y_prob.min() + 1e-9)

    y_pred = (y_prob >= 0.5).astype(int)

    acc = accuracy_score(y, y_pred)
    prec = precision_score(y, y_pred, zero_division=0)
    rec = recall_score(y, y_pred, zero_division=0)
    f1 = f1_score(y, y_pred, zero_division=0)
    try:
        auc = roc_auc_score(y, y_prob)
    except ValueError:
        auc = np.nan
    try:
        ap = average_precision_score(y, y_prob)
    except ValueError:
        ap = np.nan

    cm = confusion_matrix(y, y_pred, labels=[0, 1])
    tn, fp, fn, tp = cm.ravel() if cm.size == 4 else (0, 0, 0, 0)

    final_model = model.fit(X, y)

    metrics = {
        "Model": model_name,
        "Accuracy": acc,
        "Precision": prec,
        "Recall": rec,
        "F1-Score": f1,
        "ROC-AUC": auc,
        "PR-AUC": ap,
        "TP": int(tp),
        "FP": int(fp),
        "FN": int(fn),
        "TN": int(tn),
    }

    return metrics, final_model, y_prob, y_pred


def evaluate_regression_model(model_name: str, model, X: pd.DataFrame, y: pd.Series, cv):
    """回帰モデルの交差検証評価"""
    y_pred = cross_val_predict(model, X, y, cv=cv)

    mae = mean_absolute_error(y, y_pred)
    mse = mean_squared_error(y, y_pred)
    rmse = np.sqrt(mse)
    r2 = r2_score(y, y_pred)

    final_model = model.fit(X, y)

    metrics = {
        "Model": model_name,
        "R2-Score": r2,
        "RMSE": rmse,
        "MAE": mae,
        "MSE": mse,
    }

    return metrics, final_model, y_pred


def extract_feature_importances(models_dict: dict, feature_cols: list[str]) -> pd.DataFrame:
    """学習済みモデルから特徴量重要度・係数を抽出"""
    fi_data = {"Feature": feature_cols}
    for name, model in models_dict.items():
        estimator = model.steps[-1][1] if hasattr(model, "steps") else model
        
        if hasattr(estimator, "feature_importances_"):
            fi_data[name] = estimator.feature_importances_
        elif hasattr(estimator, "coef_"):
            coef = estimator.coef_
            fi_data[name] = np.abs(coef.ravel()) if coef.ndim > 1 else np.abs(coef)
        else:
            fi_data[name] = [np.nan] * len(feature_cols)

    df_fi = pd.DataFrame(fi_data)
    num_cols = [c for c in df_fi.columns if c != "Feature"]
    df_fi["Average_Importance"] = df_fi[num_cols].mean(axis=1)
    return df_fi.sort_values(by="Average_Importance", ascending=False).reset_index(drop=True)


def format_excel_report(excel_path: Path):
    """生成されたExcelレポートにヘッダー装飾や列幅・数値フォーマットを自動適用"""
    wb = load_workbook(excel_path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = True

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if ws.max_row >= 2 and ws.max_column >= 1:
            ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            tbl_name = f"Tbl_{ws.title[:15]}".replace(" ", "_")
            try:
                table = Table(displayName=tbl_name, ref=ref)
                table.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                ws.add_table(table)
            except ValueError:
                pass

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            header_val = str(ws.cell(1, col_idx).value or "")
            max_len = len(header_val)
            for row_idx in range(2, ws.max_row + 1):
                val = ws.cell(row_idx, col_idx).value
                if isinstance(val, (int, float)):
                    ws.cell(row_idx, col_idx).number_format = "0.0000" if isinstance(val, float) else "0"
                max_len = max(max_len, len(str(val or "")))
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 35)

    wb.save(excel_path)


def run_training(
    input_dataset: Path,
    report_dir: Path,
    model_dir: Path,
    task_type: str = TASK_TYPE,
):
    print("=" * 60)
    print(f"ML Model Training Pipeline Started - Task: {task_type.upper()}")
    print(f"Input Dataset : {input_dataset}")
    print("=" * 60)

    report_dir.mkdir(parents=True, exist_ok=True)
    model_dir.mkdir(parents=True, exist_ok=True)

    # 1. データ読み込み
    if not input_dataset.exists():
        raise FileNotFoundError(f"Input dataset not found: {input_dataset}")

    if input_dataset.suffix == ".xlsx":
        df = pd.read_excel(input_dataset)
    else:
        df = pd.read_csv(input_dataset)

    # 2. 正解ラベルの動的付与
    target_series, target_definition = define_target_label(df, task_type=task_type)
    df["target"] = target_series
    print(f"Target Label Defined: {target_definition}")

    # 特徴量列の抽出と変換
    available_features = [c for c in FEATURE_COLS if c in df.columns]
    if not available_features:
        available_features = [c for c in df.select_dtypes(include=[np.number]).columns if c not in ["source_index", "target"]]

    print(f"Using {len(available_features)} feature columns: {available_features}")

    X = df[available_features].replace([np.inf, -np.inf], np.nan).apply(pd.to_numeric, errors="coerce")
    y = pd.to_numeric(df["target"], errors="coerce")

    # 欠損行の除外
    valid_mask = X.notna().all(axis=1) & y.notna()
    X_clean = X.loc[valid_mask].reset_index(drop=True)
    y_clean = y.loc[valid_mask].reset_index(drop=True)
    df_clean = df.loc[valid_mask].reset_index(drop=True)

    print(f"Total rows: {len(df)}, Valid training rows: {len(X_clean)}")

    if len(X_clean) == 0:
        raise ValueError("有効な学習データ行が 0 件です。入力データとターゲット定義を確認してください。")

    if task_type == "classification":
        pos_cnt = int(y_clean.sum())
        neg_cnt = len(y_clean) - pos_cnt
        print(f"Class distribution -> Positive (1): {pos_cnt}, Negative (0): {neg_cnt}")

    # 3. モデル学習と評価
    metrics_list = []
    trained_models = {}
    predictions_dict = {
        "global_request_id": df_clean.get("global_request_id", df_clean.index),
        "依頼No.": df_clean.get("依頼No.", np.nan),
        "SID": df_clean.get("SID", np.nan),
        "target_actual": y_clean,
    }

    if task_type == "classification":
        models = get_classification_models()
        cv = StratifiedKFold(n_splits=N_SPLITS, shuffle=True, random_state=RANDOM_STATE)

        for name, model in models.items():
            print(f"Training & Evaluating Classifier: {name}...")
            metrics, final_model, y_prob, y_pred = evaluate_classification_model(name, model, X_clean, y_clean, cv)
            metrics_list.append(metrics)
            trained_models[name] = final_model
            predictions_dict[f"{name}_prob"] = y_prob
            predictions_dict[f"{name}_pred"] = y_pred

        df_metrics = pd.DataFrame(metrics_list).sort_values(by=["F1-Score", "ROC-AUC"], ascending=False).reset_index(drop=True)
        best_model_name = df_metrics.iloc[0]["Model"]
        print(f"\n★ Best Classification Model: {best_model_name} (F1: {df_metrics.iloc[0]['F1-Score']:.4f}, AUC: {df_metrics.iloc[0]['ROC-AUC']:.4f})")

    else:
        models = get_regression_models()
        cv = KFold(n_splits=N_SPLITS, shuffle=True, random_state=RANDOM_STATE)

        for name, model in models.items():
            print(f"Training & Evaluating Regressor: {name}...")
            metrics, final_model, y_pred = evaluate_regression_model(name, model, X_clean, y_clean, cv)
            metrics_list.append(metrics)
            trained_models[name] = final_model
            predictions_dict[f"{name}_pred"] = y_pred

        df_metrics = pd.DataFrame(metrics_list).sort_values(by=["R2-Score"], ascending=False).reset_index(drop=True)
        best_model_name = df_metrics.iloc[0]["Model"]
        print(f"\n★ Best Regression Model: {best_model_name} (R2: {df_metrics.iloc[0]['R2-Score']:.4f}, RMSE: {df_metrics.iloc[0]['RMSE']:.4f})")

    # 4. 特徴量重要度の抽出
    df_fi = extract_feature_importances(trained_models, available_features)

    # 5. 最良モデルの保存 (.joblib)
    best_model = trained_models[best_model_name]
    model_save_path = model_dir / f"best_model_{task_type}_{best_model_name}.joblib"
    package = {
        "model": best_model,
        "model_name": best_model_name,
        "task_type": task_type,
        "target_definition": target_definition,
        "feature_cols": available_features,
        "metrics": df_metrics.to_dict(orient="records"),
        "trained_at": datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M:%S"),
        "versions": {
            "python": platform.python_version(),
            "sklearn": sklearn.__version__,
            "pandas": pd.__version__,
            "numpy": np.__version__,
        },
    }
    joblib.dump(package, model_save_path, compress=3)
    print(f"Saved best model to: {model_save_path}")

    # 6. サマリーレポート (Excel & JSON) の出力
    excel_report_path = report_dir / f"model_summary_report_{task_type}.xlsx"
    df_preds = pd.DataFrame(predictions_dict)

    with pd.ExcelWriter(excel_report_path, engine="openpyxl") as writer:
        df_metrics.to_excel(writer, sheet_name="Model_Comparison", index=False)
        df_fi.to_excel(writer, sheet_name="Feature_Importances", index=False)
        df_preds.to_excel(writer, sheet_name="CV_Predictions", index=False)
        
        summary_info = pd.DataFrame([
            {"項目": "実行日時", "値": datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M:%S")},
            {"項目": "タスク種別", "値": task_type.upper()},
            {"項目": "正解ラベル定義", "値": target_definition},
            {"項目": "入力データセット", "値": str(input_dataset)},
            {"項目": "全行数", "値": len(df)},
            {"項目": "学習使用行数", "値": len(X_clean)},
            {"項目": "使用特徴量数", "値": len(available_features)},
            {"項目": "最優秀モデル", "値": best_model_name},
            {"項目": "保存モデルパス", "値": str(model_save_path)},
        ])
        summary_info.to_excel(writer, sheet_name="Run_Summary", index=False)

    format_excel_report(excel_report_path)
    print(f"Saved Excel summary report to: {excel_report_path}")

    # JSON形式サマリー保存 (正解ラベル定義 target_definition を追加)
    json_report_path = report_dir / f"summary_report_{task_type}.json"
    report_json = {
        "task_type": task_type,
        "target_definition": target_definition,
        "executed_at": datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M:%S"),
        "dataset": str(input_dataset),
        "n_samples": len(X_clean),
        "features": available_features,
        "best_model": best_model_name,
        "metrics_summary": df_metrics.to_dict(orient="records"),
        "top_features": df_fi.head(10)[["Feature", "Average_Importance"]].to_dict(orient="records"),
    }
    json_report_path.write_text(json.dumps(report_json, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Saved JSON summary report to: {json_report_path}")

    print("=" * 60)
    print("Training and Evaluation Pipeline Completed Successfully!")
    print("=" * 60)
    return df_metrics, model_save_path, excel_report_path


def main():
    parser = argparse.ArgumentParser(description="Train and evaluate ML models (Classification or Regression).")
    parser.add_argument("--input-dataset", default=str(DEFAULT_INPUT_DATASET), help=f"Path to input dataset Excel/CSV. (Default: {DEFAULT_INPUT_DATASET})")
    parser.add_argument("--report-dir", default=str(DEFAULT_REPORT_DIR), help=f"Directory for summary reports. (Default: {DEFAULT_REPORT_DIR})")
    parser.add_argument("--model-dir", default=str(DEFAULT_MODEL_DIR), help=f"Directory to save models. (Default: {DEFAULT_MODEL_DIR})")
    parser.add_argument("--task-type", default=TASK_TYPE, choices=["classification", "regression"], help=f"Task type: 'classification' or 'regression'. (Default: {TASK_TYPE})")
    args = parser.parse_args()

    run_training(
        input_dataset=Path(args.input_dataset),
        report_dir=Path(args.report_dir),
        model_dir=Path(args.model_dir),
        task_type=args.task_type,
    )


if __name__ == "__main__":
    main()
