#!/usr/bin/env python3
"""
Predict & Validate Script for Trained ML Models.

Loads a saved model (.joblib) and evaluates / predicts on unknown new dataset (.xlsx / .csv).
"""

import argparse
import logging
from datetime import datetime, timezone
from pathlib import Path

import joblib
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

PROJECT_ROOT = Path(__file__).resolve().parent.parent

logger = logging.getLogger(__name__)

# ==========================================================
# 設定エリア (パスを変更したい場合はここを書き換えてください)
# ==========================================================

def parse_path(path_str: str) -> Path:
    """Windows等の絶対パスをコピペしても安全に読み込めるようにする"""
    s = path_str.strip('\'"').replace("\\", "/").replace("¥", "/")
    p = Path(s)
    if not p.is_absolute():
        return PROJECT_ROOT / p
    return p

# 1. 検証・予測対象の未知データセット (Excel / CSV)
DEFAULT_INPUT_DATASET = parse_path(r'data/260806_検証用/260806_ML_dataset.xlsx')

# 2. 読み込む学習済みモデル (.joblib)
DEFAULT_MODEL_FILE = parse_path(r'models/trained/260806_TAT1_ml_models/best_model_regression_RandomForestRegressor.joblib')

# 3. 予測結果の出力Excelファイルパス
DEFAULT_OUTPUT_EXCEL = parse_path(r'reports/validation/regression_prediction_results.xlsx')

# ==========================================================


def format_excel_report(excel_path: Path):
    """生成されたExcelレポートにヘッダー装飾や列幅・数値フォーマットを自動適用"""
    wb = load_workbook(excel_path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = True

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if ws.max_row >= 2 and ws.max_column >= 1:
            ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            tbl_name = f"Tbl_{ws.title[:15]}".replace(" ", "_")
            try:
                table = Table(displayName=tbl_name, ref=ref)
                table.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                ws.add_table(table)
            except Exception as e:  # noqa: BLE001
                logger.warning(f"Failed to add table to worksheet {ws.title}: {e}")

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            header_val = str(ws.cell(1, col_idx).value or "")
            max_len = len(header_val)
            for row_idx in range(2, ws.max_row + 1):
                val = ws.cell(row_idx, col_idx).value
                if isinstance(val, (int, float)):
                    ws.cell(row_idx, col_idx).number_format = "0.0000" if isinstance(val, float) else "0"
                max_len = max(max_len, len(str(val or "")))
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 35)

    wb.save(excel_path)


def predict_dataset(
    input_path: Path,
    model_path: Path,
    output_excel: Path,
):
    print("=" * 60)
    print("Unknown Data Prediction & Validation Started")
    print(f"Model File    : {model_path}")
    print(f"Input Dataset : {input_path}")
    print("=" * 60)

    if not model_path.exists():
        raise FileNotFoundError(f"Model file not found: {model_path}")
    if not input_path.exists():
        raise FileNotFoundError(f"Input dataset file not found: {input_path}")

    # 1. モデルパッケージの読み込み
    pkg = joblib.load(model_path)
    model = pkg["model"]
    model_name = pkg.get("model_name", "UnknownModel")
    task_type = pkg.get("task_type", "classification")
    feature_cols = pkg.get("feature_cols", [])
    target_definition = pkg.get("target_definition", "N/A")

    print(f"Loaded Model: {model_name} (Task: {task_type.upper()})")
    print(f"Target Definition: {target_definition}")
    print(f"Required Features ({len(feature_cols)}): {feature_cols}")

    # 2. 未知データの読み込み
    if input_path.suffix.lower() in [".xlsx", ".xlsm"]:
        df = pd.read_excel(input_path)
    else:
        df = pd.read_csv(input_path)

    # 3. 特徴量の抽出と整合性チェック
    missing_cols = [c for c in feature_cols if c not in df.columns]
    if missing_cols:
        raise ValueError(f"入力データに必要な特徴量カラムが含まれていません: {missing_cols}")

    X = df[feature_cols].replace([np.inf, -np.inf], np.nan).apply(pd.to_numeric, errors="coerce")
    valid_mask = X.notna().all(axis=1)

    print(f"Input total rows: {len(df)}, Scored valid rows: {int(valid_mask.sum())}")

    df_valid = df.loc[valid_mask].copy().reset_index(drop=True)
    X_valid = X.loc[valid_mask].reset_index(drop=True)

    # 4. 推論の実行
    output_df = pd.DataFrame()
    for col in ["source_index", "source_file", "global_request_id", "依頼No.", "SID", "属性", "SampleType"]:
        if col in df_valid.columns:
            output_df[col] = df_valid[col]

    if task_type == "classification":
        if hasattr(model, "predict_proba"):
            probs = model.predict_proba(X_valid)[:, 1]
            output_df["predicted_probability"] = probs
            output_df["predicted_class"] = (probs >= 0.5).astype(int)
        else:
            preds = model.predict(X_valid)
            output_df["predicted_class"] = preds
    else:
        preds = model.predict(X_valid)
        output_df["predicted_value"] = preds

    # 5. 出力Excelの保存
    output_excel.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
        output_df.to_excel(writer, sheet_name="Predictions", index=False)
        
        info_df = pd.DataFrame([
            {"項目": "推論実行日時", "値": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")},
            {"項目": "使用モデル", "値": model_name},
            {"項目": "タスク種別", "値": task_type.upper()},
            {"項目": "正解ラベル定義", "値": target_definition},
            {"項目": "モデルファイル", "値": str(model_path)},
            {"項目": "入力データセット", "値": str(input_path)},
            {"項目": "全行数", "値": len(df)},
            {"項目": "推論成功行数", "値": int(valid_mask.sum())},
        ])
        info_df.to_excel(writer, sheet_name="Run_Info", index=False)

    format_excel_report(output_excel)
    print(f"Saved prediction results to: {output_excel}")
    print("=" * 60)
    print("Prediction Completed Successfully!")
    print("=" * 60)


def main():
    parser = argparse.ArgumentParser(description="Predict and validate using a trained ML model.")
    parser.add_argument("--input", default=str(DEFAULT_INPUT_DATASET), help=f"Input Excel/CSV dataset. (Default: {DEFAULT_INPUT_DATASET})")
    parser.add_argument("--model", default=str(DEFAULT_MODEL_FILE), help=f"Trained .joblib model file. (Default: {DEFAULT_MODEL_FILE})")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT_EXCEL), help=f"Output Excel prediction path. (Default: {DEFAULT_OUTPUT_EXCEL})")
    args = parser.parse_args()

    predict_dataset(
        input_path=Path(args.input),
        model_path=Path(args.model),
        output_excel=Path(args.output),
    )


if __name__ == "__main__":
    main()
