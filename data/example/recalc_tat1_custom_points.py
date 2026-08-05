#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TAT1 custom photometric-point concentration recalculation.

Purpose
-------
Recalculate TAT1 concentration from time-course absorbance data for these
photometric point ranges:
  2-8, 4-10, 6-12, 8-14, 10-16, 12-18, 14-20

Input files, by default, are expected in the same directory as this script:
  - measurement.csv
  - profile.csv
  - metadata.json  optional, used only for Summary sheet

Output:
  - TAT1_custom_points_recalc.xlsx
  - TAT1_custom_points_recalc_recalc_matrix.csv
  - TAT1_custom_points_recalc_rates.csv
  - TAT1_custom_points_recalc_cal_rates.csv

Calibration concentrations used:
  Cal0-Cal5 = [0.0, 4.7, 13.8, 28.8, 61.9, 126.0] ng/mL
"""

from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


TARGET_ITEM = "TAT1"

PT_PATTERNS = [
    ("2-8", 2, 8),
    ("4-10", 4, 10),
    ("6-12", 6, 12),
    ("8-14", 8, 14),
    ("10-16", 10, 16),
    ("12-18", 12, 18),
    ("14-20", 14, 20),
]

PT_TIMES = {
    1: 0.0, 2: 9.2, 3: 18.0, 4: 27.2, 5: 36.0, 6: 45.2, 7: 54.0,
    8: 63.2, 9: 72.0, 10: 81.2, 11: 90.0, 12: 99.2, 13: 108.0,
    14: 117.2, 15: 126.0, 16: 135.2, 17: 144.0, 18: 153.2,
    19: 162.0, 20: 171.2, 21: 180.0,
}

STD_CONCS = [0.0, 4.7, 13.8, 28.8, 61.9, 126.0]


def clean_text(value):
    """Clean embedded nulls and accidental double quoting from CSV exports."""
    if pd.isna(value):
        return value
    text = str(value).replace("\x00", "").strip()
    for _ in range(2):
        if len(text) >= 2 and text[0] == '"' and text[-1] == '"':
            text = text[1:-1]
    return text.replace('""', '"')


def load_inputs(input_dir: Path):
    measurement_path = input_dir / "measurement.csv"
    profile_path = input_dir / "profile.csv"
    metadata_path = input_dir / "metadata.json"

    if not measurement_path.exists():
        raise FileNotFoundError(f"measurement.csv not found: {measurement_path}")
    if not profile_path.exists():
        raise FileNotFoundError(f"profile.csv not found: {profile_path}")

    meas_df = pd.read_csv(measurement_path, dtype=str, encoding="utf-8-sig", engine="python")
    profile_df = pd.read_csv(profile_path, dtype=str, encoding="utf-8-sig", engine="python")

    for df in (meas_df, profile_df):
        for col in df.columns:
            df[col] = df[col].map(clean_text)

    if TARGET_ITEM not in meas_df.columns:
        raise ValueError(f"{TARGET_ITEM} column is missing from measurement.csv")

    for col in ["source_index", TARGET_ITEM]:
        if col in meas_df.columns:
            meas_df[col] = pd.to_numeric(meas_df[col], errors="coerce")

    for col in ["source_index", "項目No.", "測光ﾎﾟｰﾄ", "処理値", "時間", "吸光度"]:
        if col in profile_df.columns:
            profile_df[col] = pd.to_numeric(profile_df[col], errors="coerce")

    if "source_index" in meas_df.columns:
        meas_df["source_index"] = meas_df["source_index"].astype("Int64")
    if "source_index" in profile_df.columns:
        profile_df["source_index"] = profile_df["source_index"].astype("Int64")

    # Rebuild global_request_id to avoid issues from quoted request IDs.
    if {"source_index", "依頼No."}.issubset(meas_df.columns):
        meas_df["global_request_id"] = (
            meas_df["source_index"].astype(str).str.replace("<NA>", "", regex=False)
            + "_"
            + meas_df["依頼No."].astype(str)
        )
    if {"source_index", "依頼No."}.issubset(profile_df.columns):
        profile_df["global_request_id"] = (
            profile_df["source_index"].astype(str).str.replace("<NA>", "", regex=False)
            + "_"
            + profile_df["依頼No."].astype(str)
        )

    metadata = {}
    if metadata_path.exists():
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))

    return meas_df, profile_df, metadata


def classify_sample_type(value) -> str:
    text = str(value).strip()
    upper = text.upper()
    if text in ["", "None", "nan", "NaN"]:
        return "OTHER"
    if "CAL" in upper or "キャリブ" in text or "SALINE" in upper or "生食" in text:
        return "CAL"
    if "CTRL" in upper or "CONTROL" in upper or "CONT" in upper or "コントロール" in text or "パネル" in text:
        return "QC"
    return "SAMPLE"


def calc_rate_mabs_min(time_values, absorbance_values, pt_start: int, pt_end: int) -> float:
    """Calculate absorbance slope in mAbs/min between two photometric points."""
    if pt_start not in PT_TIMES or pt_end not in PT_TIMES:
        return np.nan

    times = np.asarray(time_values, dtype=float)
    absorbance = np.asarray(absorbance_values, dtype=float)
    if len(times) == 0 or len(times) != len(absorbance):
        return np.nan

    t_start = PT_TIMES[pt_start]
    t_end = PT_TIMES[pt_end]
    idx_start = int(np.argmin(np.abs(times - t_start)))
    idx_end = int(np.argmin(np.abs(times - t_end)))

    delta_t = times[idx_end] - times[idx_start]
    if delta_t <= 0:
        return np.nan

    return float(((absorbance[idx_end] - absorbance[idx_start]) * 0.1) / (delta_t / 60.0))


def representative(values) -> float:
    """Representative value: n>=3 median of first 3, n=2 mean, n=1 raw."""
    arr = [float(v) for v in values if not pd.isna(v)]
    if len(arr) >= 3:
        return float(np.median(arr[:3]))
    if len(arr) == 2:
        return float(np.mean(arr))
    if len(arr) == 1:
        return float(arr[0])
    return np.nan


def interpolate_or_extrapolate(rate, cal_rates, cal_concs) -> float:
    """Piecewise linear interpolation with endpoint extrapolation."""
    if pd.isna(rate) or cal_rates is None or len(cal_rates) < 2:
        return np.nan

    r = np.asarray(cal_rates, dtype=float)
    c = np.asarray(cal_concs, dtype=float)
    order = np.argsort(r)
    r = r[order]
    c = c[order]
    r, unique_idx = np.unique(r, return_index=True)
    c = c[unique_idx]

    if len(r) < 2:
        return np.nan

    if r[0] <= rate <= r[-1]:
        return float(np.interp(rate, r, c))

    if rate < r[0]:
        denom = r[1] - r[0]
        if denom == 0:
            return float(c[0])
        return float(c[0] + (c[1] - c[0]) / denom * (rate - r[0]))

    denom = r[-1] - r[-2]
    if denom == 0:
        return float(c[-1])
    return float(c[-1] + (c[-1] - c[-2]) / denom * (rate - r[-1]))


def cal_level_from_attr(attr):
    """Map attribute text to Cal level where possible."""
    text = str(attr).strip()
    upper = text.upper()
    if "SALINE" in upper or "生食" in text:
        return 0
    match = re.search(r"CAL\s*[-_ ]?\s*(\d+)", upper)
    if match:
        level = int(match.group(1))
        return level if 0 <= level <= 5 else None
    return None


def build_rates(profile_df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    profile = profile_df[profile_df["項目名"].astype(str) == TARGET_ITEM].copy()
    profile["時間"] = pd.to_numeric(profile["時間"], errors="coerce")
    profile["吸光度"] = pd.to_numeric(profile["吸光度"], errors="coerce")
    profile = profile.dropna(subset=["時間", "吸光度"])

    group_cols = ["source_index", "source_file", "global_request_id", "依頼No.", "項目名"]
    grouped_curves = []
    for keys, group in profile.groupby(group_cols, dropna=False):
        base = dict(zip(group_cols, keys))
        sorted_group = group.sort_values("時間")
        grouped_curves.append((
            base,
            sorted_group["時間"].to_numpy(float),
            sorted_group["吸光度"].to_numpy(float),
        ))

    rates_by_pattern = {}
    for pattern_name, pt_start, pt_end in PT_PATTERNS:
        rows = []
        for base, times, absorbance in grouped_curves:
            row = base.copy()
            row["測光Ptパターン"] = pattern_name
            row["Rate_mAbs_min"] = calc_rate_mabs_min(times, absorbance, pt_start, pt_end)
            rows.append(row)
        rates_by_pattern[pattern_name] = pd.DataFrame(rows)

    return rates_by_pattern


def build_calibration_curves(meas_df: pd.DataFrame, rates_by_pattern: dict[str, pd.DataFrame]):
    """Build calibration curves by photometric pattern and source_index."""
    cal_rows = []
    for _, row in meas_df.iterrows():
        request_no = str(row.get("依頼No.", "")).strip()
        c_match = re.match(r"^C(\d+)$", request_no)
        attr_level = cal_level_from_attr(row.get("属性", ""))

        if not c_match and attr_level is None:
            continue

        raw_value = pd.to_numeric(row.get(TARGET_ITEM, np.nan), errors="coerce")
        if pd.isna(raw_value):
            continue

        cal_rows.append({
            "source_index": row.get("source_index", "GLOBAL"),
            "source_file": row.get("source_file", ""),
            "global_request_id": row.get("global_request_id", ""),
            "依頼No.": request_no,
            "属性": row.get("属性", ""),
            "項目名": TARGET_ITEM,
            "装置生データ": raw_value,
            "cal_source_type": "C_ID" if c_match else "ATTR",
            "c_num": int(c_match.group(1)) if c_match else np.nan,
            "cal_level_attr": attr_level if attr_level is not None else np.nan,
        })

    cal_df = pd.DataFrame(cal_rows)
    curves_by_pattern_source = {}
    summary_rows = []

    if cal_df.empty:
        return curves_by_pattern_source, pd.DataFrame()

    for pattern_name, rate_df in rates_by_pattern.items():
        curves_by_pattern_source[pattern_name] = {}
        joined = cal_df.merge(
            rate_df[["source_index", "global_request_id", "項目名", "Rate_mAbs_min"]],
            on=["source_index", "global_request_id", "項目名"],
            how="left",
        )
        # For calibration only, if time-course rate is missing, use instrument raw value.
        joined["cal_rate"] = pd.to_numeric(joined["Rate_mAbs_min"], errors="coerce").fillna(
            pd.to_numeric(joined["装置生データ"], errors="coerce")
        )

        for source in list(joined["source_index"].dropna().unique()) + ["GLOBAL"]:
            source_df = joined if source == "GLOBAL" else joined[joined["source_index"].astype(str) == str(source)]
            if source_df.empty:
                continue

            buckets = {i: [] for i in range(len(STD_CONCS))}

            # C001, C002... style: every 3 records become one Cal level.
            c_sub = source_df[source_df["cal_source_type"] == "C_ID"].dropna(subset=["c_num", "cal_rate"]).sort_values("c_num")
            if len(c_sub) >= 2:
                values = c_sub["cal_rate"].astype(float).tolist()
                idx = 0
                level = 0
                while idx < len(values) and level < len(STD_CONCS):
                    buckets[level].append(representative(values[idx:idx + 3]))
                    idx += 3
                    level += 1

            # Attribute style: Saline/Cal-1/Cal-2...
            attr_sub = source_df[source_df["cal_level_attr"].notna()].dropna(subset=["cal_rate"])
            for level, group in attr_sub.groupby("cal_level_attr"):
                level = int(level)
                if 0 <= level < len(STD_CONCS):
                    buckets[level].append(representative(group["cal_rate"].tolist()))

            cal_rates = []
            cal_concs = []
            for level, conc in enumerate(STD_CONCS):
                rate_value = representative(buckets[level])
                if not pd.isna(rate_value):
                    cal_rates.append(rate_value)
                    cal_concs.append(conc)

            if len(cal_rates) >= 2:
                cal_rates = np.array(cal_rates, dtype=float)
                cal_concs = np.array(cal_concs, dtype=float)
                curves_by_pattern_source[pattern_name][source] = (cal_rates, cal_concs)

                record = {
                    "測光Ptパターン": pattern_name,
                    "source_index": source,
                    "項目名": TARGET_ITEM,
                    "有効Cal点数": len(cal_rates),
                }
                for i, (rv, cv) in enumerate(zip(cal_rates, cal_concs)):
                    record[f"CalPoint{i}_濃度_ng_mL"] = cv
                    record[f"CalPoint{i}_Rate_mAbs_min"] = rv
                summary_rows.append(record)

    return curves_by_pattern_source, pd.DataFrame(summary_rows)


def find_curve(curves, pattern_name, source_index):
    if pattern_name in curves:
        if source_index in curves[pattern_name]:
            cal_rates, cal_concs = curves[pattern_name][source_index]
            return cal_rates, cal_concs, f"source_index={source_index}"
        if "GLOBAL" in curves[pattern_name]:
            cal_rates, cal_concs = curves[pattern_name]["GLOBAL"]
            return cal_rates, cal_concs, "GLOBAL"
    return None, None, "NO_CURVE"


def format_workbook(path: Path):
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if worksheet.max_row >= 2 and worksheet.max_column >= 1:
            ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
            table_name = re.sub(r"[^A-Za-z0-9_]", "_", worksheet.title)[:20] + "Tbl"
            try:
                table = Table(displayName=table_name, ref=ref)
                table.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                worksheet.add_table(table)
            except Exception:
                pass

        for col_idx in range(1, worksheet.max_column + 1):
            header = str(worksheet.cell(1, col_idx).value or "")
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(len(header) + 2, 10), 34)
            for row_idx in range(2, worksheet.max_row + 1):
                value = worksheet.cell(row_idx, col_idx).value
                if isinstance(value, (int, float)):
                    worksheet.cell(row_idx, col_idx).number_format = "0.000"

    workbook.save(path)


def run(input_dir: Path, output_excel: Path):
    meas_df, profile_df, metadata = load_inputs(input_dir)
    meas_df["SampleType"] = meas_df.get("属性", "").apply(classify_sample_type)

    rates_by_pattern = build_rates(profile_df)
    curves, cal_summary_df = build_calibration_curves(meas_df, rates_by_pattern)
    rate_long_df = pd.concat(rates_by_pattern.values(), ignore_index=True) if rates_by_pattern else pd.DataFrame()

    base_cols = ["source_index", "source_file", "global_request_id", "依頼No.", "SID", "属性", "SampleType"]
    rows = []
    for _, sample in meas_df[base_cols].drop_duplicates("global_request_id").iterrows():
        global_id = str(sample["global_request_id"])
        source_index = sample.get("source_index", "GLOBAL")
        measurement_row = meas_df[meas_df["global_request_id"].astype(str) == global_id].iloc[0]

        output_row = {col: sample[col] for col in base_cols}
        output_row[f"{TARGET_ITEM}_装置生データ"] = pd.to_numeric(measurement_row.get(TARGET_ITEM, np.nan), errors="coerce")

        for pattern_name, _, _ in PT_PATTERNS:
            pattern_rates = rates_by_pattern[pattern_name]
            matched_rate = pattern_rates[
                (pattern_rates["global_request_id"].astype(str) == global_id)
                & (pattern_rates["項目名"].astype(str) == TARGET_ITEM)
            ]
            rate_value = pd.to_numeric(matched_rate.iloc[0]["Rate_mAbs_min"], errors="coerce") if not matched_rate.empty else np.nan

            cal_rates, cal_concs, curve_source = find_curve(curves, pattern_name, source_index)
            output_row[f"{TARGET_ITEM}_{pattern_name}"] = interpolate_or_extrapolate(rate_value, cal_rates, cal_concs)
            output_row[f"{TARGET_ITEM}_{pattern_name}_Rate"] = rate_value
            output_row[f"{TARGET_ITEM}_{pattern_name}_CalCurve"] = curve_source

        rows.append(output_row)

    recalc_matrix_df = pd.DataFrame(rows)

    summary_df = pd.DataFrame([
        {"項目": "実行日時", "値": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"項目": "対象項目", "値": TARGET_ITEM},
        {"項目": "測光ポイント", "値": ", ".join([p[0] for p in PT_PATTERNS])},
        {"項目": "使用標準濃度_ng_mL", "値": ", ".join(map(str, STD_CONCS))},
        {"項目": "measurement行数", "値": len(meas_df)},
        {"項目": "profile行数", "値": len(profile_df)},
        {"項目": "再計算マトリクス形状", "値": str(recalc_matrix_df.shape)},
        {"項目": "Rate一覧形状", "値": str(rate_long_df.shape)},
        {"項目": "キャリブRate形状", "値": str(cal_summary_df.shape)},
        {"項目": "入力CSV", "値": ", ".join(metadata.get("source_csv_files", []))},
    ])

    output_excel.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
        recalc_matrix_df.to_excel(writer, sheet_name="TAT1_再計算", index=False)
        rate_long_df.to_excel(writer, sheet_name="Rate一覧", index=False)
        cal_summary_df.to_excel(writer, sheet_name="キャリブRate", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

    format_workbook(output_excel)

    recalc_matrix_df.to_csv(output_excel.with_name(output_excel.stem + "_recalc_matrix.csv"), index=False, encoding="utf-8-sig")
    rate_long_df.to_csv(output_excel.with_name(output_excel.stem + "_rates.csv"), index=False, encoding="utf-8-sig")
    cal_summary_df.to_csv(output_excel.with_name(output_excel.stem + "_cal_rates.csv"), index=False, encoding="utf-8-sig")

    return {
        "output_excel": str(output_excel),
        "measurement_shape": tuple(meas_df.shape),
        "profile_shape": tuple(profile_df.shape),
        "recalc_matrix_shape": tuple(recalc_matrix_df.shape),
        "rate_long_shape": tuple(rate_long_df.shape),
        "cal_rate_shape": tuple(cal_summary_df.shape),
        "target_item": TARGET_ITEM,
        "pt_patterns": [p[0] for p in PT_PATTERNS],
        "std_concs": STD_CONCS,
    }


def main():
    parser = argparse.ArgumentParser(description="Recalculate TAT1 concentration from custom photometric points.")
    parser.add_argument("--input-dir", default=".", help="Directory containing measurement.csv, profile.csv, and optionally metadata.json")
    parser.add_argument("--output-excel", default="TAT1_custom_points_recalc.xlsx", help="Output Excel file path")
    args = parser.parse_args()

    result = run(Path(args.input_dir), Path(args.output_excel))
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
